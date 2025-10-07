"""
CallSession ViewSet for aiMediaGateway

This module contains ViewSets for the CallSession model with tenant filtering
and comprehensive call session management.
"""

from rest_framework import viewsets, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from django.utils import timezone
from django.http import HttpResponse
from datetime import timedelta
import csv
import io

from core.models import CallSession, UserProfile, AudioRecording
from core.serializers import (
    CallSessionSerializer,
    CallSessionDetailSerializer,
    CallSessionCreateSerializer,
    LiveCallStatusSerializer
)
from core.junie_codes.permissions import TenantResourcePermission
from core.junie_codes.filters import CallSessionFilterSet


class CallSessionViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing call sessions with tenant filtering.
    
    Provides CRUD operations for call session management with proper
    tenant isolation and role-based permissions.
    """
    
    queryset = CallSession.objects.all()
    serializer_class = CallSessionSerializer
    permission_classes = [IsAuthenticated, TenantResourcePermission]
    
    # Advanced filtering and search configuration
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_class = CallSessionFilterSet
    search_fields = ['caller_id', 'caller_name', 'dialed_number', 'channel_name', 'session_id']
    ordering_fields = ['start_time', 'end_time', 'created_at', 'updated_at', 'caller_id', 'dialed_number']
    ordering = ['-start_time']
    
    def get_queryset(self):
        """
        Filter queryset based on tenant context and user permissions.
        
        Returns:
            QuerySet: Filtered call session queryset
        """
        queryset = super().get_queryset()
        
        # Add related data for efficiency
        queryset = queryset.select_related('tenant').prefetch_related('audiorecording_set')
        
        # Filter by tenant context (set by middleware)
        if hasattr(self.request, 'tenant') and self.request.tenant:
            queryset = queryset.filter(tenant=self.request.tenant)
        else:
            # Fallback to user's tenant if no tenant context
            try:
                user_profile = UserProfile.objects.get(user=self.request.user)
                queryset = queryset.filter(tenant=user_profile.tenant)
            except UserProfile.DoesNotExist:
                return queryset.none()
        
        # Additional filtering based on user role
        if hasattr(self.request, 'user') and self.request.user.is_authenticated:
            try:
                user_profile = UserProfile.objects.get(user=self.request.user)
                
                # Viewers can see all sessions in their tenant
                # Operators and admins can see all sessions
                # No additional filtering needed for now
                
            except UserProfile.DoesNotExist:
                return queryset.none()
        
        return queryset
    
    def get_serializer_class(self):
        """
        Return appropriate serializer based on action.
        
        Returns:
            Serializer class: Appropriate serializer for the action
        """
        if self.action == 'create':
            return CallSessionCreateSerializer
        elif self.action == 'retrieve':
            return CallSessionDetailSerializer
        elif self.action in ['live_status', 'update_status']:
            return LiveCallStatusSerializer
        return CallSessionSerializer
    
    def perform_create(self, serializer):
        """
        Create a call session with proper tenant assignment.
        
        Args:
            serializer: CallSessionCreateSerializer instance
        """
        # Set tenant from request context
        if hasattr(self.request, 'tenant') and self.request.tenant:
            serializer.save(tenant=self.request.tenant)
        else:
            # Fallback to user's tenant
            try:
                user_profile = UserProfile.objects.get(user=self.request.user)
                serializer.save(tenant=user_profile.tenant)
            except UserProfile.DoesNotExist:
                # This should not happen due to permission checks
                raise ValueError("No tenant context available")
    
    def perform_update(self, serializer):
        """
        Update call session with audit logging.
        
        Args:
            serializer: CallSessionSerializer instance
        """
        old_instance = self.get_object()
        new_instance = serializer.save()
        
        # Log status changes
        if old_instance.status != new_instance.status:
            print(f"Call session {new_instance.session_id} status changed from "
                  f"{old_instance.status} to {new_instance.status}")
    
    @action(detail=False, methods=['get'])
    def active(self, request):
        """
        Get all active call sessions for the current tenant.
        
        Args:
            request: HTTP request
            
        Returns:
            Response: List of active call sessions
        """
        active_statuses = ['active', 'ringing', 'connecting']
        queryset = self.get_queryset().filter(status__in=active_statuses)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def recent(self, request):
        """
        Get recent call sessions (last 24 hours).
        
        Args:
            request: HTTP request
            
        Returns:
            Response: List of recent call sessions
        """
        since = timezone.now() - timedelta(hours=24)
        queryset = self.get_queryset().filter(start_time__gte=since)
        
        # Paginate the results
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def start_recording(self, request, pk=None):
        """
        Start recording for a call session.
        
        Args:
            request: HTTP request
            pk: Call session primary key
            
        Returns:
            Response: Success or error message
        """
        call_session = self.get_object()
        
        if not call_session.recording_enabled:
            return Response(
                {'message': 'Recording is not enabled for this session'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if call_session.status not in ['active', 'ringing']:
            return Response(
                {'message': 'Can only start recording for active calls'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check if recording is already active
        existing_recording = AudioRecording.objects.filter(
            call_session=call_session,
            # Add status field check when available
        ).first()
        
        if existing_recording:
            return Response(
                {'message': 'Recording is already active for this session'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # TODO: Implement actual recording start logic
        # This would typically involve:
        # 1. Creating an AudioRecording instance
        # 2. Starting the actual recording process via Asterisk ARI
        # 3. Setting up file paths and metadata
        
        return Response({
            'message': 'Recording started successfully',
            'session_id': call_session.session_id
        })
    
    @action(detail=True, methods=['post'])
    def stop_recording(self, request, pk=None):
        """
        Stop recording for a call session.
        
        Args:
            request: HTTP request
            pk: Call session primary key
            
        Returns:
            Response: Success or error message
        """
        call_session = self.get_object()
        
        # Find active recording
        active_recording = AudioRecording.objects.filter(
            call_session=call_session,
            # Add status field check when available
        ).first()
        
        if not active_recording:
            return Response(
                {'message': 'No active recording found for this session'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # TODO: Implement actual recording stop logic
        # This would typically involve:
        # 1. Stopping the recording process via Asterisk ARI
        # 2. Updating the AudioRecording instance with final metadata
        # 3. Triggering transcription if enabled
        
        return Response({
            'message': 'Recording stopped successfully',
            'session_id': call_session.session_id,
            'recording_id': active_recording.id
        })
    
    @action(detail=True, methods=['post'])
    def end_call(self, request, pk=None):
        """
        End a call session.
        
        Args:
            request: HTTP request
            pk: Call session primary key
            
        Returns:
            Response: Success or error message
        """
        call_session = self.get_object()
        
        if call_session.status in ['completed', 'failed']:
            return Response(
                {'message': 'Call is already ended'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Update call session
        call_session.status = 'completed'
        call_session.end_time = timezone.now()
        call_session.save()
        
        # TODO: Implement actual call termination via Asterisk ARI
        
        return Response({
            'message': 'Call ended successfully',
            'session_id': call_session.session_id,
            'duration': call_session.duration if hasattr(call_session, 'duration') else None
        })
    
    @action(detail=True, methods=['get', 'post'])
    def live_status(self, request, pk=None):
        """
        Get or update live status of a call session.
        
        Args:
            request: HTTP request
            pk: Call session primary key
            
        Returns:
            Response: Live status data
        """
        call_session = self.get_object()
        
        if request.method == 'GET':
            # Return current live status
            duration_seconds = None
            if call_session.start_time:
                if call_session.end_time:
                    duration_seconds = int((call_session.end_time - call_session.start_time).total_seconds())
                elif call_session.status in ['active', 'ringing']:
                    duration_seconds = int((timezone.now() - call_session.start_time).total_seconds())
            
            status_data = {
                'session_id': call_session.session_id,
                'status': call_session.status,
                'caller_id': call_session.caller_id,
                'dialed_number': call_session.dialed_number,
                'duration_seconds': duration_seconds,
                'recording_status': 'active' if call_session.recording_enabled else 'disabled',
                'participant_count': 2,  # Default for now
                'quality_metrics': call_session.metadata.get('quality_metrics') if call_session.metadata else None,
                'timestamp': timezone.now(),
                'tenant_id': call_session.tenant.id,
                'tenant_slug': call_session.tenant.slug,
            }
            
            serializer = LiveCallStatusSerializer(data=status_data)
            serializer.is_valid(raise_exception=True)
            return Response(serializer.data)
        
        else:  # POST - Update status
            serializer = LiveCallStatusSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            
            # Update call session with new status data
            validated_data = serializer.validated_data
            
            if 'status' in validated_data:
                call_session.status = validated_data['status']
            
            if 'quality_metrics' in validated_data and validated_data['quality_metrics']:
                if not call_session.metadata:
                    call_session.metadata = {}
                call_session.metadata['quality_metrics'] = validated_data['quality_metrics']
            
            call_session.save()
            
            return Response({
                'message': 'Live status updated successfully',
                'session_id': call_session.session_id
            })
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """
        Get call session statistics for the current tenant.
        
        Args:
            request: HTTP request
            
        Returns:
            Response: Call session statistics
        """
        queryset = self.get_queryset()
        
        # Calculate date ranges
        now = timezone.now()
        today = now.date()
        week_ago = now - timedelta(days=7)
        month_ago = now - timedelta(days=30)
        
        stats = {
            'total_sessions': queryset.count(),
            'active_sessions': queryset.filter(status__in=['active', 'ringing', 'connecting']).count(),
            'completed_sessions': queryset.filter(status='completed').count(),
            'failed_sessions': queryset.filter(status='failed').count(),
            'sessions_today': queryset.filter(start_time__date=today).count(),
            'sessions_this_week': queryset.filter(start_time__gte=week_ago).count(),
            'sessions_this_month': queryset.filter(start_time__gte=month_ago).count(),
            'inbound_sessions': queryset.filter(call_direction='inbound').count(),
            'outbound_sessions': queryset.filter(call_direction='outbound').count(),
            'recorded_sessions': queryset.filter(recording_enabled=True).count(),
            'transcribed_sessions': queryset.filter(transcription_enabled=True).count(),
        }
        
        # Calculate average call duration for completed calls
        completed_sessions = queryset.filter(
            status='completed',
            end_time__isnull=False,
            start_time__isnull=False
        )
        
        if completed_sessions.exists():
            durations = [
                (session.end_time - session.start_time).total_seconds()
                for session in completed_sessions
            ]
            stats['average_duration_seconds'] = sum(durations) / len(durations)
        else:
            stats['average_duration_seconds'] = None
        
        stats['generated_at'] = now
        
        return Response(stats)
    
    @action(detail=False, methods=['post'], url_path='bulk-create')
    def bulk_create(self, request):
        """
        Create multiple call sessions in bulk.
        
        Args:
            request: HTTP request with list of call session data
            
        Returns:
            Response: List of created call sessions or validation errors
        """
        if not isinstance(request.data, list):
            return Response(
                {'error': 'Request data must be a list of call session objects'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if len(request.data) > 100:  # Limit bulk operations
            return Response(
                {'error': 'Maximum 100 call sessions can be created in bulk'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        created_sessions = []
        errors = []
        
        for index, session_data in enumerate(request.data):
            serializer = CallSessionCreateSerializer(data=session_data)
            
            if serializer.is_valid():
                try:
                    # Set tenant from request context
                    if hasattr(request, 'tenant') and request.tenant:
                        session = serializer.save(tenant=request.tenant)
                    else:
                        # Fallback to user's tenant
                        try:
                            user_profile = UserProfile.objects.get(user=request.user)
                            session = serializer.save(tenant=user_profile.tenant)
                        except UserProfile.DoesNotExist:
                            errors.append({
                                'index': index,
                                'error': 'No tenant context available'
                            })
                            continue
                    
                    created_sessions.append({
                        'index': index,
                        'session_id': session.session_id,
                        'id': session.id
                    })
                    
                except Exception as e:
                    errors.append({
                        'index': index,
                        'error': str(e)
                    })
            else:
                errors.append({
                    'index': index,
                    'validation_errors': serializer.errors
                })
        
        response_data = {
            'created_count': len(created_sessions),
            'error_count': len(errors),
            'created_sessions': created_sessions,
            'errors': errors
        }
        
        response_status = status.HTTP_201_CREATED if created_sessions else status.HTTP_400_BAD_REQUEST
        return Response(response_data, status=response_status)
    
    @action(detail=False, methods=['patch'], url_path='bulk-update')
    def bulk_update(self, request):
        """
        Update multiple call sessions in bulk.
        
        Expected format:
        {
            "session_ids": ["session1", "session2", ...],
            "data": {"field": "value", ...}
        }
        
        Args:
            request: HTTP request with session IDs and update data
            
        Returns:
            Response: Count of updated sessions and any errors
        """
        session_ids = request.data.get('session_ids', [])
        update_data = request.data.get('data', {})
        
        if not session_ids:
            return Response(
                {'error': 'session_ids list is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not update_data:
            return Response(
                {'error': 'data object with fields to update is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if len(session_ids) > 100:  # Limit bulk operations
            return Response(
                {'error': 'Maximum 100 call sessions can be updated in bulk'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get sessions that belong to the current tenant
        queryset = self.get_queryset().filter(session_id__in=session_ids)
        
        updated_sessions = []
        errors = []
        
        for session in queryset:
            serializer = CallSessionSerializer(session, data=update_data, partial=True)
            
            if serializer.is_valid():
                try:
                    updated_session = serializer.save()
                    updated_sessions.append({
                        'session_id': updated_session.session_id,
                        'id': updated_session.id
                    })
                except Exception as e:
                    errors.append({
                        'session_id': session.session_id,
                        'error': str(e)
                    })
            else:
                errors.append({
                    'session_id': session.session_id,
                    'validation_errors': serializer.errors
                })
        
        # Check for session IDs that weren't found
        found_session_ids = [s.session_id for s in queryset]
        not_found_ids = [sid for sid in session_ids if sid not in found_session_ids]
        
        for session_id in not_found_ids:
            errors.append({
                'session_id': session_id,
                'error': 'Session not found or access denied'
            })
        
        response_data = {
            'updated_count': len(updated_sessions),
            'error_count': len(errors),
            'updated_sessions': updated_sessions,
            'errors': errors
        }
        
        return Response(response_data)
    
    @action(detail=False, methods=['delete'], url_path='bulk-delete')
    def bulk_delete(self, request):
        """
        Delete multiple call sessions in bulk.
        
        Expected format:
        {
            "session_ids": ["session1", "session2", ...]
        }
        
        Args:
            request: HTTP request with session IDs to delete
            
        Returns:
            Response: Count of deleted sessions and any errors
        """
        session_ids = request.data.get('session_ids', [])
        
        if not session_ids:
            return Response(
                {'error': 'session_ids list is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if len(session_ids) > 100:  # Limit bulk operations
            return Response(
                {'error': 'Maximum 100 call sessions can be deleted in bulk'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get sessions that belong to the current tenant
        queryset = self.get_queryset().filter(session_id__in=session_ids)
        
        # Check if any sessions are still active
        active_sessions = queryset.filter(status__in=['active', 'ringing', 'connecting'])
        if active_sessions.exists():
            return Response(
                {
                    'error': 'Cannot delete active call sessions',
                    'active_session_ids': list(active_sessions.values_list('session_id', flat=True))
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        
        deleted_sessions = []
        errors = []
        
        for session in queryset:
            try:
                session_id = session.session_id
                session.delete()
                deleted_sessions.append({
                    'session_id': session_id,
                    'id': session.id
                })
            except Exception as e:
                errors.append({
                    'session_id': session.session_id,
                    'error': str(e)
                })
        
        # Check for session IDs that weren't found
        found_session_ids = [s.session_id for s in queryset]
        not_found_ids = [sid for sid in session_ids if sid not in found_session_ids]
        
        for session_id in not_found_ids:
            errors.append({
                'session_id': session_id,
                'error': 'Session not found or access denied'
            })
        
        response_data = {
            'deleted_count': len(deleted_sessions),
            'error_count': len(errors),
            'deleted_sessions': deleted_sessions,
            'errors': errors
        }
        
        return Response(response_data)
    
    @action(detail=True, methods=['get'], url_path='recordings')
    def recordings(self, request, pk=None):
        """
        Get all audio recordings for a specific call session.
        
        Args:
            request: HTTP request
            pk: Call session primary key
            
        Returns:
            Response: List of audio recordings for the session
        """
        call_session = self.get_object()
        recordings = AudioRecording.objects.filter(call_session=call_session)
        
        # Import serializer locally to avoid circular imports
        from core.serializers import AudioRecordingSerializer
        
        serializer = AudioRecordingSerializer(recordings, many=True)
        return Response({
            'session_id': call_session.session_id,
            'recording_count': recordings.count(),
            'recordings': serializer.data
        })
    
    @action(detail=False, methods=['post'], url_path='bulk-end-calls')
    def bulk_end_calls(self, request):
        """
        End multiple active call sessions in bulk.
        
        Expected format:
        {
            "session_ids": ["session1", "session2", ...],
            "reason": "Manual termination"  # Optional
        }
        
        Args:
            request: HTTP request with session IDs to end
            
        Returns:
            Response: Count of ended calls and any errors
        """
        session_ids = request.data.get('session_ids', [])
        reason = request.data.get('reason', 'Bulk termination')
        
        if not session_ids:
            return Response(
                {'error': 'session_ids list is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if len(session_ids) > 50:  # Lower limit for call termination
            return Response(
                {'error': 'Maximum 50 call sessions can be ended in bulk'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get active sessions that belong to the current tenant
        queryset = self.get_queryset().filter(
            session_id__in=session_ids,
            status__in=['active', 'ringing', 'connecting']
        )
        
        ended_sessions = []
        errors = []
        
        for session in queryset:
            try:
                session.status = 'completed'
                session.end_time = timezone.now()
                
                # Add termination reason to metadata
                if not session.metadata:
                    session.metadata = {}
                session.metadata['termination_reason'] = reason
                session.metadata['terminated_by'] = request.user.username
                
                session.save()
                
                ended_sessions.append({
                    'session_id': session.session_id,
                    'id': session.id,
                    'end_time': session.end_time
                })
                
                # TODO: Implement actual call termination via Asterisk ARI
                
            except Exception as e:
                errors.append({
                    'session_id': session.session_id,
                    'error': str(e)
                })
        
        # Check for session IDs that weren't found or not active
        found_session_ids = [s.session_id for s in queryset]
        not_found_ids = [sid for sid in session_ids if sid not in found_session_ids]
        
        for session_id in not_found_ids:
            # Check if session exists but is not active
            existing_session = self.get_queryset().filter(session_id=session_id).first()
            if existing_session:
                errors.append({
                    'session_id': session_id,
                    'error': f'Session is not active (current status: {existing_session.status})'
                })
            else:
                errors.append({
                    'session_id': session_id,
                    'error': 'Session not found or access denied'
                })
        
        response_data = {
            'ended_count': len(ended_sessions),
            'error_count': len(errors),
            'ended_sessions': ended_sessions,
            'errors': errors
        }
        
        return Response(response_data)
    
    @action(detail=False, methods=['get'], url_path='export-csv')
    def export_csv(self, request):
        """
        Export call sessions to CSV format.
        
        Supports all the same filtering options as the list view.
        
        Args:
            request: HTTP request with optional filtering parameters
            
        Returns:
            HttpResponse: CSV file download
        """
        # Apply the same filtering as the list view
        queryset = self.filter_queryset(self.get_queryset())
        
        # Limit export size to prevent memory issues
        if queryset.count() > 10000:
            return Response(
                {'error': 'Export size too large. Please use filters to reduce the result set.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Create CSV response
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="call_sessions_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        
        # Write header row
        header = [
            'Session ID',
            'Caller ID',
            'Caller Name',
            'Dialed Number',
            'Call Direction',
            'Status',
            'Start Time',
            'End Time',
            'Duration (seconds)',
            'Recording Enabled',
            'Transcription Enabled',
            'Channel Name',
            'Tenant',
            'Created At',
            'Updated At',
        ]
        writer.writerow(header)
        
        # Write data rows
        for session in queryset:
            # Calculate duration
            duration = None
            if session.start_time and session.end_time:
                duration = int((session.end_time - session.start_time).total_seconds())
            
            row = [
                session.session_id,
                session.caller_id,
                session.caller_name or '',
                session.dialed_number,
                session.call_direction,
                session.status,
                session.start_time.isoformat() if session.start_time else '',
                session.end_time.isoformat() if session.end_time else '',
                duration or '',
                session.recording_enabled,
                session.transcription_enabled,
                session.channel_name or '',
                session.tenant.name if session.tenant else '',
                session.created_at.isoformat(),
                session.updated_at.isoformat(),
            ]
            writer.writerow(row)
        
        return response
    
    @action(detail=False, methods=['post'], url_path='import-csv')
    def import_csv(self, request):
        """
        Import call sessions from CSV file with validation.
        
        Expected CSV format matches the export format.
        
        Args:
            request: HTTP request with CSV file in request.FILES
            
        Returns:
            Response: Import results with success/error counts
        """
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No file provided. Please upload a CSV file.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        csv_file = request.FILES['file']
        
        # Validate file type
        if not csv_file.name.endswith('.csv'):
            return Response(
                {'error': 'Invalid file format. Please upload a CSV file.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check file size (limit to 10MB)
        if csv_file.size > 10 * 1024 * 1024:
            return Response(
                {'error': 'File too large. Maximum size is 10MB.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Decode file content
            file_data = csv_file.read().decode('utf-8')
            io_string = io.StringIO(file_data)
            reader = csv.DictReader(io_string)
            
            imported_count = 0
            errors = []
            skipped_count = 0
            
            # Get tenant context
            if hasattr(request, 'tenant') and request.tenant:
                tenant = request.tenant
            else:
                try:
                    user_profile = UserProfile.objects.get(user=request.user)
                    tenant = user_profile.tenant
                except UserProfile.DoesNotExist:
                    return Response(
                        {'error': 'No tenant context available'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            
            for row_num, row in enumerate(reader, 2):  # Start from row 2 (after header)
                try:
                    # Check if session already exists
                    session_id = row.get('Session ID', '').strip()
                    if session_id and CallSession.objects.filter(
                        session_id=session_id,
                        tenant=tenant
                    ).exists():
                        skipped_count += 1
                        continue
                    
                    # Create CallSession instance
                    session_data = {
                        'session_id': session_id or f'import_{timezone.now().timestamp()}_{row_num}',
                        'caller_id': row.get('Caller ID', '').strip(),
                        'caller_name': row.get('Caller Name', '').strip() or None,
                        'dialed_number': row.get('Dialed Number', '').strip(),
                        'call_direction': row.get('Call Direction', 'inbound').strip(),
                        'status': row.get('Status', 'completed').strip(),
                        'recording_enabled': row.get('Recording Enabled', 'False').strip().lower() == 'true',
                        'transcription_enabled': row.get('Transcription Enabled', 'False').strip().lower() == 'true',
                        'channel_name': row.get('Channel Name', '').strip() or None,
                        'tenant': tenant,
                    }
                    
                    # Parse dates
                    start_time_str = row.get('Start Time', '').strip()
                    if start_time_str:
                        try:
                            session_data['start_time'] = timezone.datetime.fromisoformat(
                                start_time_str.replace('Z', '+00:00')
                            )
                        except ValueError:
                            pass
                    
                    end_time_str = row.get('End Time', '').strip()
                    if end_time_str:
                        try:
                            session_data['end_time'] = timezone.datetime.fromisoformat(
                                end_time_str.replace('Z', '+00:00')
                            )
                        except ValueError:
                            pass
                    
                    # Validate using serializer
                    serializer = CallSessionCreateSerializer(data=session_data)
                    if serializer.is_valid():
                        serializer.save()
                        imported_count += 1
                    else:
                        errors.append({
                            'row': row_num,
                            'errors': serializer.errors
                        })
                
                except Exception as e:
                    errors.append({
                        'row': row_num,
                        'error': str(e)
                    })
            
            return Response({
                'imported_count': imported_count,
                'skipped_count': skipped_count,
                'error_count': len(errors),
                'errors': errors[:100],  # Limit error details
                'message': f'Import completed. {imported_count} sessions imported, {skipped_count} skipped, {len(errors)} errors.'
            })
            
        except Exception as e:
            return Response(
                {'error': f'Failed to process CSV file: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        """
        Export call sessions to Excel format with formatting.
        
        Supports all the same filtering options as the list view.
        Creates a formatted Excel file with headers, styling, and data validation.
        
        Args:
            request: HTTP request with optional filtering parameters
            
        Returns:
            HttpResponse: Excel file download
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return Response(
                {'error': 'Excel export requires openpyxl package. Please install it.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        # Apply the same filtering as the list view
        queryset = self.filter_queryset(self.get_queryset())
        
        # Limit export size to prevent memory issues
        if queryset.count() > 10000:
            return Response(
                {'error': 'Export size too large. Please use filters to reduce the result set.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Call Sessions"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = [
            'Session ID',
            'Caller ID',
            'Caller Name',
            'Dialed Number',
            'Call Direction',
            'Status',
            'Start Time',
            'End Time',
            'Duration (seconds)',
            'Recording Enabled',
            'Transcription Enabled',
            'Channel Name',
            'Tenant',
            'Created At',
            'Updated At',
        ]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Write data
        row_num = 2
        for session in queryset:
            # Calculate duration
            duration = None
            if session.start_time and session.end_time:
                duration = int((session.end_time - session.start_time).total_seconds())
            
            data = [
                session.session_id,
                session.caller_id,
                session.caller_name or '',
                session.dialed_number,
                session.call_direction,
                session.status,
                session.start_time.isoformat() if session.start_time else '',
                session.end_time.isoformat() if session.end_time else '',
                duration,
                'Yes' if session.recording_enabled else 'No',
                'Yes' if session.transcription_enabled else 'No',
                session.channel_name or '',
                session.tenant.name if session.tenant else '',
                session.created_at.isoformat(),
                session.updated_at.isoformat(),
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border
                
                # Apply specific formatting
                if col_num in [7, 8, 14, 15]:  # Date columns
                    cell.alignment = Alignment(horizontal="center")
                elif col_num == 9:  # Duration column
                    cell.alignment = Alignment(horizontal="right")
                elif col_num in [10, 11]:  # Boolean columns
                    cell.alignment = Alignment(horizontal="center")
                    if value == 'Yes':
                        cell.fill = PatternFill(start_color="D4F3D0", end_color="D4F3D0", fill_type="solid")
            
            row_num += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    cell_length = len(str(cell.value)) if cell.value else 0
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)  # Cap at 30 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add summary information in a separate sheet
        summary_ws = wb.create_sheet(title="Summary")
        
        # Summary data
        total_sessions = queryset.count()
        active_sessions = queryset.filter(status__in=['active', 'ringing', 'connecting']).count()
        completed_sessions = queryset.filter(status='completed').count()
        failed_sessions = queryset.filter(status__in=['failed', 'timeout', 'error']).count()
        recorded_sessions = queryset.filter(recording_enabled=True).count()
        
        summary_data = [
            ['Export Summary', ''],
            ['Generated At', timezone.now().isoformat()],
            ['Total Sessions', total_sessions],
            ['Active Sessions', active_sessions],
            ['Completed Sessions', completed_sessions],
            ['Failed Sessions', failed_sessions],
            ['Recorded Sessions', recorded_sessions],
            ['Recording Rate', f"{(recorded_sessions/total_sessions*100):.1f}%" if total_sessions > 0 else "0%"],
        ]
        
        for row_num, (label, value) in enumerate(summary_data, 1):
            summary_ws.cell(row=row_num, column=1, value=label).font = Font(bold=True)
            summary_ws.cell(row=row_num, column=2, value=value)
        
        # Auto-adjust summary sheet columns
        for column in summary_ws.columns:
            max_length = max(len(str(cell.value)) for cell in column if cell.value)
            column_letter = get_column_letter(column[0].column)
            summary_ws.column_dimensions[column_letter].width = max_length + 2
        
        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Create response
        response = HttpResponse(
            excel_buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="call_sessions_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        return response
